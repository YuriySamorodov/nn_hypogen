import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from zipfile import BadZipFile

from src.corpus import is_lfs_pointer
from src.settings import Settings, get_settings


SIZE_CLASS_RE = re.compile(r"^[+-]?\d+(?:\s*\+\s*\d+)?(?:\s*мкм)?$", re.IGNORECASE)
SKIP_LOSS_ROWS = {"Итого (проверка)", "Потери (расписать)", "Свободный слот"}


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"#REF!", "#DIV/0!", "#VALUE!"} else text


def _to_float(value: object) -> float | None:
    text = _clean_text(value).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _fact_text(
    *,
    example: str,
    subject: str,
    metric: str,
    value: float,
    unit: str,
    source_path: str,
    row: int,
) -> str:
    return (
        f"{example}: {subject}; показатель: {metric}; значение: {value:g} {unit}. "
        f"Источник: {source_path}, строка {row}."
    )


def _add_node(nodes: dict[str, dict[str, Any]], node_id: str, label: str, name: str) -> None:
    nodes.setdefault(node_id, {"id": node_id, "label": label, "name": name})


def _add_edge(
    edges: list[dict[str, str]],
    source: str,
    relation: str,
    target: str,
) -> None:
    edge = {"source": source, "relation": relation, "target": target}
    if edge not in edges:
        edges.append(edge)


def _add_fact(
    facts: list[dict[str, Any]],
    *,
    example: str,
    source_path: Path,
    sheet: str,
    row: int,
    subject: str,
    metric: str,
    value: float | None,
    unit: str,
    element: str | None = None,
) -> None:
    if value is None:
        return
    fact_id = f"FACT-{len(facts) + 1:05d}"
    metric_name = f"{metric} {element}".strip() if element else metric
    facts.append(
        {
            "id": fact_id,
            "example": example,
            "source_id": f"GRAPH-{source_path.parent.name}-{source_path.stem}",
            "source_path": str(source_path),
            "sheet": sheet,
            "row": row,
            "subject": subject,
            "metric": metric,
            "element": element,
            "value": value,
            "unit": unit,
            "text": _fact_text(
                example=example,
                subject=subject,
                metric=metric_name,
                value=value,
                unit=unit,
                source_path=str(source_path),
                row=row,
            ),
        }
    )


def _is_size_class(value: str) -> bool:
    normalized = value.replace(" ", "")
    return bool(SIZE_CLASS_RE.match(normalized)) and any(ch.isdigit() for ch in normalized)


def _parse_xlsx(
    path: Path,
    nodes: dict[str, dict[str, Any]],
    edges: list[dict[str, str]],
    facts: list[dict[str, Any]],
) -> None:
    example = path.parent.name
    example_id = f"example:{example}"
    _add_node(nodes, example_id, "Example", example)

    workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        mode: str | None = None
        current_tail = "отвальные хвосты"
        current_size_class: str | None = None

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            cells = list(row)
            first = _clean_text(cells[1] if len(cells) > 1 and not _clean_text(cells[0]) else cells[0])
            if not first:
                continue

            lowered = first.lower()
            if first == "Класс крупности, мкм":
                mode = "size_distribution"
                continue

            if "доля потерь" in " ".join(_clean_text(cell).lower() for cell in cells):
                current_size_class = first.replace(" мкм", "").strip()
                mode = "loss_forms"
                continue

            if "хвост" in lowered and _to_float(cells[2] if len(cells) > 2 else None) is not None:
                current_tail = first
                tail_id = f"tail:{example}:{current_tail}"
                _add_node(nodes, tail_id, "TailProduct", current_tail)
                _add_edge(edges, example_id, "HAS_TAIL_PRODUCT", tail_id)
                subject = f"{current_tail}"
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="масса хвостов",
                    value=_to_float(cells[2] if len(cells) > 2 else None),
                    unit="СМТ",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="содержание",
                    element="Элемент 28",
                    value=_to_float(cells[3] if len(cells) > 3 else None),
                    unit="%",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="содержание",
                    element="Элемент 29",
                    value=_to_float(cells[5] if len(cells) > 5 else None),
                    unit="%",
                )
                continue

            if lowered.startswith("доля ") and "хвост" in lowered:
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=first,
                    metric="доля",
                    value=_to_float(cells[2] if len(cells) > 2 else cells[1] if len(cells) > 1 else None),
                    unit="%",
                )
                continue

            if mode == "size_distribution" and _is_size_class(first):
                current_size_class = first
                size_id = f"size:{example}:{current_tail}:{first}"
                tail_id = f"tail:{example}:{current_tail}"
                _add_node(nodes, size_id, "SizeClass", first)
                _add_node(nodes, tail_id, "TailProduct", current_tail)
                _add_edge(edges, tail_id, "HAS_SIZE_CLASS", size_id)
                subject = f"{current_tail} / класс крупности {first}"
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="доля класса",
                    value=_to_float(cells[2] if len(cells) > 2 else None),
                    unit="%",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="доля металла в классе",
                    element="Элемент 28",
                    value=_to_float(cells[3] if len(cells) > 3 else None),
                    unit="%",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="доля металла в классе",
                    element="Элемент 29",
                    value=_to_float(cells[5] if len(cells) > 5 else None),
                    unit="%",
                )
                continue

            if mode == "loss_forms" and current_size_class and first not in SKIP_LOSS_ROWS:
                subject = f"{current_tail} / класс {current_size_class} / {first}"
                form_id = f"form:{example}:{current_tail}:{current_size_class}:{first}"
                size_id = f"size:{example}:{current_tail}:{current_size_class}"
                _add_node(nodes, form_id, "LossForm", first)
                _add_node(nodes, size_id, "SizeClass", current_size_class)
                _add_edge(edges, size_id, "HAS_LOSS_FORM", form_id)
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="доля потерь",
                    element="Элемент 28",
                    value=_to_float(cells[2] if len(cells) > 2 else None),
                    unit="%",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="потери",
                    element="Элемент 28",
                    value=_to_float(cells[3] if len(cells) > 3 else None),
                    unit="т",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="доля потерь",
                    element="Элемент 29",
                    value=_to_float(cells[4] if len(cells) > 4 else None),
                    unit="%",
                )
                _add_fact(
                    facts,
                    example=example,
                    source_path=path,
                    sheet=worksheet.title,
                    row=row_index,
                    subject=subject,
                    metric="потери",
                    element="Элемент 29",
                    value=_to_float(cells[5] if len(cells) > 5 else None),
                    unit="т",
                )


def build_knowledge_graph(
    settings: Settings | None = None,
    *,
    task_dir: Path | None = None,
    output_path: Path | None = None,
) -> dict[str, Any]:
    settings = settings or get_settings()
    task_dir = task_dir or Path(settings.task_data_dir)
    output_path = output_path or Path(settings.knowledge_graph_path)

    nodes: dict[str, dict[str, Any]] = {}
    edges: list[dict[str, str]] = []
    facts: list[dict[str, Any]] = []

    for path in sorted(task_dir.glob("Пример */*.xlsx")):
        if is_lfs_pointer(path):
            continue
        try:
            _parse_xlsx(path, nodes, edges, facts)
        except BadZipFile:
            continue

    if not facts and output_path.exists():
        return json.loads(output_path.read_text(encoding="utf-8"))

    graph = {
        "nodes": sorted(nodes.values(), key=lambda item: item["id"]),
        "edges": edges,
        "facts": facts,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(graph, ensure_ascii=False, indent=2), encoding="utf-8")
    return graph


def main() -> None:
    graph = build_knowledge_graph()
    print(
        f"Built knowledge graph: {len(graph['nodes'])} nodes, "
        f"{len(graph['edges'])} edges, {len(graph['facts'])} facts."
    )


if __name__ == "__main__":
    main()

